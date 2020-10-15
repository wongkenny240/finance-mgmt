"""
Author: KCKW
Description: fetch finance data for pandas dataframe and write to Excel
"""

import pandas_datareader.data as web
import pandas as pd
from openpyxl import load_workbook


if __name__ == '__main__':
    # Define the instruments to download. We would like to see Apple, Microsoft and the S&P500 index.
    tickers = ['AAPL', 'AMZN']
    # We would like all available data from 01/01/2000 until 12/31/2016.
    start_date = '1980-01-01'
    end_date = '2020-10-05'

    for ticker in tickers:
        # User pandas_reader.data.DataReader to load the desired data. As simple as that.
        df = web.DataReader(ticker, 'yahoo', start_date, end_date)
        df = df.reset_index()
        df['Date'] = df['Date'].dt.date
        print(df.columns)

        # write to excel
        writer = pd.ExcelWriter("stock.xlsx", engine='openpyxl')
        book = load_workbook("stock.xlsx")

        # delete sheet
        delete_sheet = [ticker]
        for sheetName in book.sheetnames:
            if sheetName in delete_sheet:
                del book[sheetName]
        book.save("stock.xlsx")

        # set write parameter
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        # output to excel
        df.to_excel(writer,
                          sheet_name=ticker, index=False)

        writer.save()
        writer.close()
